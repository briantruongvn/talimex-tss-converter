"""
Bước 3: Copy dữ liệu từ Input và điền X cho Products
- Đọc file input để lấy data từ row 10 trở đi
- Mở file Step2 và copy data theo column mapping
- Điền X vào các cột product
- Xuất file: {input_name}-Step3.xlsx
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment

INPUT_FOLDER = "input"
OUTPUT_FOLDER = "output"

# Column mapping: Input column -> Output column
COLUMN_MAPPING = {
    2: 17,  # B -> Q
    4: 2,   # D -> B
    5: 3,   # E -> C
    6: 4,   # F -> D
    7: 5,   # G -> E
    8: 6,   # H -> F
    9: 7,   # I -> G
    10: 8,  # J -> H
    11: 9,  # K -> I
    12: 10, # L -> J
    14: 11, # N -> K
    15: 12, # O -> L
}


def get_input_files() -> list[Path]:
    """Lấy tất cả file Excel từ folder input (bỏ qua file tạm ~$)"""
    input_path = Path(INPUT_FOLDER)
    files = []
    for pattern in ["*.xlsx", "*.xls"]:
        for f in input_path.glob(pattern):
            if not f.name.startswith("~$"):
                files.append(f)
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file Excel trong folder {INPUT_FOLDER}")
    return files


def get_sheets_except_material_code(wb):
    """Lấy tất cả sheet trừ 'material code' (case insensitive)"""
    sheets = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() != "material code":
            sheets.append(wb[sheet_name])
    return sheets


def copy_data(input_ws, output_ws):
    """Copy data từ input sang output theo mapping"""
    data_start_row = 10  # Data bắt đầu từ row 10 trong input
    output_start_row = 11  # Output bắt đầu từ row 11

    row_count = 0
    for input_row in range(data_start_row, input_ws.max_row + 1):
        # Kiểm tra row có data không (check cột đầu tiên trong mapping)
        has_data = False
        for input_col in COLUMN_MAPPING.keys():
            if input_ws.cell(row=input_row, column=input_col).value:
                has_data = True
                break

        if not has_data:
            continue

        output_row = output_start_row + row_count

        for input_col, output_col in COLUMN_MAPPING.items():
            value = input_ws.cell(row=input_row, column=input_col).value
            output_ws.cell(row=output_row, column=output_col, value=value)

        row_count += 1

    return row_count


def count_products(output_ws):
    """Đếm số cột product (từ column R onwards có giá trị ở row 10)"""
    count = 0
    col = 18  # Column R = 18
    while output_ws.cell(row=10, column=col).value:
        count += 1
        col += 1
    return count


def fill_product_marks(output_ws, num_products, data_rows):
    """Điền X vào các cột product từ row 11 đến hết data"""
    start_col = 18  # Column R = 18
    start_row = 11

    center_alignment = Alignment(horizontal='center', vertical='center')

    for row in range(start_row, start_row + data_rows):
        for i in range(num_products):
            col = start_col + i
            cell = output_ws.cell(row=row, column=col, value="X")
            cell.alignment = center_alignment


def main():
    input_files = get_input_files()

    print(f"Tìm thấy {len(input_files)} file trong folder input")

    for input_file in input_files:
        input_name = input_file.stem
        step2_file = Path(OUTPUT_FOLDER) / f"{input_name}-Step2.xlsx"

        if not step2_file.exists():
            print(f"  ⚠ Bỏ qua {input_file.name}: Chưa có file Step2")
            continue

        # Đọc file input
        input_wb = load_workbook(input_file)
        input_sheets = get_sheets_except_material_code(input_wb)

        if not input_sheets:
            print(f"  ⚠ Bỏ qua {input_file.name}: Không tìm thấy sheet phù hợp")
            continue

        # Mở file Step2
        output_wb = load_workbook(step2_file)
        output_ws = output_wb.active

        # Copy data từ tất cả sheets (trừ material code)
        total_rows = 0
        for input_ws in input_sheets:
            rows_copied = copy_data(input_ws, output_ws)
            total_rows += rows_copied

        # Đếm số products và điền X
        num_products = count_products(output_ws)
        if num_products > 0 and total_rows > 0:
            fill_product_marks(output_ws, num_products, total_rows)

        # Lưu file Step3
        output_filename = f"{input_name}-Step3.xlsx"
        output_path = Path(OUTPUT_FOLDER) / output_filename
        output_wb.save(output_path)

        print(f"  {input_file.name} → {output_filename}")
        print(f"    Copied {total_rows} rows, {num_products} products marked with X")

    print(f"\nHoàn thành!")


if __name__ == "__main__":
    main()
