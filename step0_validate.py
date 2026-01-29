"""
Bước 0: Validate format file input trước khi chạy pipeline
- Kiểm tra header row (row 9) có đúng thứ tự các cột không
- Báo lỗi chi tiết nếu format không đúng
- Chỉ cho phép tiếp tục pipeline nếu tất cả file input hợp lệ
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from dataclasses import dataclass


INPUT_FOLDER = "input"
HEADER_ROW = 9  # Row 9 chứa header

# Expected headers: column index (1-based) -> expected text (lowercase, partial match)
EXPECTED_HEADERS = {
    2: "material",              # B - MATERIAL
    4: "general type",          # D - General Type Component(Type) Process Type
    5: "sub-type",              # E - Sub-Type Component Identity Process Name
    6: "material designation",  # F
    7: "material distributor",  # G
    8: "producer",              # H
    9: "material type",         # I - Material Type in Process
    10: "document type",        # J
    11: "requirement source",   # K - Requirement Source/ TED
    12: "sub-type",             # L
    14: "details",              # N - Details of requirement
    15: "test requirement",     # O - Test requirement Result
}


@dataclass
class ValidationError:
    """Chi tiết lỗi validation cho một cột"""
    column: int
    column_letter: str
    expected: str
    actual: str


@dataclass
class ValidationResult:
    """Kết quả validation cho một file"""
    file_path: Path
    is_valid: bool
    errors: list[ValidationError]


def get_column_letter(col_idx: int) -> str:
    """Chuyển đổi column index (1-based) sang letter (A, B, C...)"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def get_input_files() -> list[Path]:
    """Lấy tất cả file Excel từ folder input (bỏ qua file tạm ~$)"""
    input_path = Path(INPUT_FOLDER)
    files = list(input_path.glob("*.xlsx")) + list(input_path.glob("*.xls"))
    # Bỏ qua file tạm của Excel (bắt đầu bằng ~$)
    files = [f for f in files if not f.name.startswith("~$")]
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file Excel trong folder {INPUT_FOLDER}")
    return sorted(files)


def validate_file(file_path: Path) -> ValidationResult:
    """
    Validate một file Excel
    - Kiểm tra header row có đúng format không
    - So sánh case-insensitive, partial match
    """
    errors = []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        for col_idx, expected_text in EXPECTED_HEADERS.items():
            cell_value = ws.cell(row=HEADER_ROW, column=col_idx).value
            actual_text = str(cell_value).strip() if cell_value else ""

            # Partial match, case-insensitive
            if expected_text.lower() not in actual_text.lower():
                errors.append(ValidationError(
                    column=col_idx,
                    column_letter=get_column_letter(col_idx),
                    expected=expected_text,
                    actual=actual_text if actual_text else "(empty)"
                ))

        wb.close()

    except Exception as e:
        errors.append(ValidationError(
            column=0,
            column_letter="-",
            expected="Readable Excel file",
            actual=f"Error: {str(e)}"
        ))

    return ValidationResult(
        file_path=file_path,
        is_valid=len(errors) == 0,
        errors=errors
    )


def print_validation_result(result: ValidationResult) -> None:
    """In kết quả validation cho một file"""
    if result.is_valid:
        print(f"  ✓ {result.file_path.name} - Valid")
    else:
        print(f"  ✗ {result.file_path.name} - Invalid")
        for error in result.errors:
            print(f"      Column {error.column_letter} ({error.column}): "
                  f"expected '{error.expected}', got '{error.actual}'")


def main() -> int:
    """
    Main function
    Returns:
        0 if all files valid
        1 if any file invalid
    """
    print("=" * 60)
    print("Step 0: Validating input file format")
    print("=" * 60)

    try:
        input_files = get_input_files()
    except FileNotFoundError as e:
        print(f"Error: {e}")
        return 1

    print(f"\nFound {len(input_files)} file(s) in '{INPUT_FOLDER}/' folder")
    print(f"Checking header row {HEADER_ROW}...\n")

    results = [validate_file(f) for f in input_files]

    valid_count = sum(1 for r in results if r.is_valid)
    invalid_count = len(results) - valid_count

    for result in results:
        print_validation_result(result)

    print()
    print("-" * 60)

    if invalid_count == 0:
        print(f"✓ All {valid_count} file(s) are valid. Ready to proceed.")
        return 0
    else:
        print(f"✗ {invalid_count} file(s) have invalid format.")
        print("  Please fix the format before running the pipeline.")
        return 1


if __name__ == "__main__":
    sys.exit(main())
