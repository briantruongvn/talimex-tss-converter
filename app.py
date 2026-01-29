"""
TALIMEX Internal TSS Converter - Streamlit App
Convert TALIMEX Internal TSS files to Standard TSS format.
"""

import streamlit as st
import zipfile
import io
from pathlib import Path
from openpyxl import load_workbook

# Import UI toolkit components
from streamlit_ui_toolkit import get_tss_17column_template

# Import validation logic from step0
from step0_validate import EXPECTED_HEADERS, HEADER_ROW, ValidationError, ValidationResult

# Page config
st.set_page_config(
    page_title="TALIMEX TSS Converter",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# Custom CSS
st.markdown("""
<style>
    /* Hide Streamlit default elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stDeployButton {display: none;}

    /* Main container */
    .main-header {
        text-align: center;
        padding: 1rem 0 0.5rem 0;
    }

    .main-title {
        font-size: 2.5rem;
        font-weight: 700;
        color: #1f2937;
        margin-bottom: 0.5rem;
    }

    .main-subtitle {
        font-size: 1rem;
        color: #6b7280;
        font-weight: 400;
        margin-bottom: 1rem;
    }

    .divider {
        border-top: 1px solid #e5e7eb;
        margin: 1rem 0 2rem 0;
    }

    /* Upload section */
    .upload-container {
        border: 2px dashed #d1d5db;
        border-radius: 12px;
        padding: 2rem;
        text-align: center;
        background: #fafafa;
        margin-bottom: 1rem;
    }

    .upload-title {
        font-size: 1.25rem;
        font-weight: 600;
        color: #374151;
        margin-bottom: 0.5rem;
    }

    .upload-subtitle {
        font-size: 0.95rem;
        color: #6b7280;
    }

    /* Validation results */
    .file-valid {
        color: #059669;
        font-weight: 500;
    }

    .file-invalid {
        color: #dc2626;
        font-weight: 500;
    }

    /* Center button */
    div.stButton > button {
        background: #3b82f6;
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.75rem 2rem;
        font-weight: 500;
        font-size: 1rem;
    }

    div.stButton > button:hover {
        background: #2563eb;
    }
</style>
""", unsafe_allow_html=True)


def validate_file_content(file_bytes: bytes, filename: str) -> ValidationResult:
    """Validate file content without saving to disk"""
    errors = []

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        for col_idx, expected_text in EXPECTED_HEADERS.items():
            cell_value = ws.cell(row=HEADER_ROW, column=col_idx).value
            actual_text = str(cell_value).strip() if cell_value else ""

            if expected_text.lower() not in actual_text.lower():
                errors.append(ValidationError(
                    column=col_idx,
                    column_letter=get_column_letter(col_idx),
                    expected=expected_text,
                    actual=actual_text if actual_text else "(empty)"
                ))

        wb.close()

    except Exception as e:
        errors.append(ValidationError(
            column=0,
            column_letter="-",
            expected="Readable Excel file",
            actual=f"Error: {str(e)}"
        ))

    return ValidationResult(
        file_path=Path(filename),
        is_valid=len(errors) == 0,
        errors=errors
    )


def get_column_letter(col_idx: int) -> str:
    """Convert column index (1-based) to letter"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def get_sheets_except_material_code(wb):
    """Get all sheets except 'material code'"""
    sheets = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() != "material code":
            sheets.append(wb[sheet_name])
    return sheets


def find_product_info(ws):
    """Find Product name and Article number in first 3 rows"""
    product_names = []
    article_numbers = []

    for row in range(1, 4):
        for col in range(1, min(ws.max_column + 1, 20)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_str = str(cell_value).lower().strip()
                if "product name" in cell_str:
                    for offset in range(1, 4):
                        value = ws.cell(row=row, column=col + offset).value
                        if value:
                            product_names = [p.strip() for p in str(value).split('\n') if p.strip()]
                            break
                elif "article number" in cell_str:
                    for offset in range(1, 4):
                        value = ws.cell(row=row, column=col + offset).value
                        if value:
                            article_numbers = [str(a).strip() for a in str(value).split('\n') if str(a).strip()]
                            break

    return product_names, article_numbers


def process_file(input_bytes: bytes, input_filename: str, progress_callback=None) -> bytes:
    """Process a single file through all pipeline steps."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter as openpyxl_get_column_letter

    COLUMN_MAPPING = {
        2: 17, 4: 2, 5: 3, 6: 4, 7: 5, 8: 6,
        9: 7, 10: 8, 11: 9, 12: 10, 14: 11, 15: 12,
    }

    DATA_START_ROW = 11

    # Step 1: Create template
    if progress_callback:
        progress_callback(1, "Creating template...")

    template = get_tss_17column_template()

    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = "TSS Data"

    for col_idx, col_config in enumerate(template.columns, 1):
        cell = output_ws.cell(template.header_row, col_idx, col_config.name)
        cell.font = Font(bold=col_config.font_bold, color=col_config.font_color)
        cell.fill = PatternFill(start_color=col_config.bg_color, end_color=col_config.bg_color, fill_type="solid")
        cell.alignment = Alignment(horizontal=col_config.horizontal_align, vertical=col_config.vertical_align, wrap_text=col_config.wrap_text)
        col_letter = openpyxl_get_column_letter(col_idx)
        output_ws.column_dimensions[col_letter].width = col_config.width

    if template.freeze_panes:
        output_ws.freeze_panes = template.freeze_panes

    # Step 2: Fill product info
    if progress_callback:
        progress_callback(2, "Filling product info...")

    input_wb = load_workbook(io.BytesIO(input_bytes))
    input_sheets = get_sheets_except_material_code(input_wb)

    all_product_names = []
    all_article_numbers = []

    for input_ws in input_sheets:
        product_names, article_numbers = find_product_info(input_ws)
        if product_names:
            all_product_names.extend(product_names)
        if article_numbers:
            all_article_numbers.extend(article_numbers)

    start_col = 18
    num_products = max(len(all_product_names), len(all_article_numbers))

    header_font = Font(bold=False)
    header_alignment = Alignment(textRotation=90, vertical='center', horizontal='center', wrap_text=True)
    article_alignment = Alignment(vertical='center', horizontal='center')
    peach_fill = PatternFill(start_color="FFFCD5B4", end_color="FFFCD5B4", fill_type="solid")

    for i in range(num_products):
        col = start_col + i
        col_letter = output_ws.cell(row=1, column=col).column_letter

        name = all_product_names[i] if i < len(all_product_names) else ""
        article = all_article_numbers[i] if i < len(all_article_numbers) else ""

        article_cell = output_ws.cell(row=10, column=col, value=article)
        article_cell.alignment = article_alignment
        article_cell.font = header_font
        article_cell.fill = peach_fill

        output_ws.merge_cells(start_row=1, start_column=col, end_row=9, end_column=col)
        name_cell = output_ws.cell(row=1, column=col, value=name)
        name_cell.alignment = header_alignment
        name_cell.font = header_font
        name_cell.fill = peach_fill

        for row in range(1, 10):
            cell = output_ws.cell(row=row, column=col)
            cell.fill = peach_fill

        article_len = len(str(article)) if article else 0
        width = max(article_len + 2, 10)
        output_ws.column_dimensions[col_letter].width = width

    # Step 3: Copy data
    if progress_callback:
        progress_callback(3, "Copying data...")

    data_start_row = 10
    output_start_row = 11
    total_rows = 0

    for input_ws in input_sheets:
        for input_row in range(data_start_row, input_ws.max_row + 1):
            has_data = False
            for input_col in COLUMN_MAPPING.keys():
                if input_ws.cell(row=input_row, column=input_col).value:
                    has_data = True
                    break

            if not has_data:
                continue

            output_row = output_start_row + total_rows

            for input_col, output_col in COLUMN_MAPPING.items():
                value = input_ws.cell(row=input_row, column=input_col).value
                output_ws.cell(row=output_row, column=output_col, value=value)

            total_rows += 1

    center_alignment = Alignment(horizontal='center', vertical='center')
    for row in range(output_start_row, output_start_row + total_rows):
        for i in range(num_products):
            col = start_col + i
            cell = output_ws.cell(row=row, column=col, value="X")
            cell.alignment = center_alignment

    input_wb.close()

    # Step 4: Cleanup
    if progress_callback:
        progress_callback(4, "Cleaning up data...")

    for row in range(DATA_START_ROW, output_ws.max_row + 1):
        h_val = output_ws.cell(row=row, column=8).value
        if h_val:
            h_lower = str(h_val).lower().strip()
            if h_lower not in ('test report', 'tr'):
                output_ws.cell(row=row, column=11).value = None

    for row in range(DATA_START_ROW, output_ws.max_row + 1):
        q_val = output_ws.cell(row=row, column=17).value
        if q_val:
            q_lower = str(q_val).lower()
            if 'article' in q_lower or 'art' in q_lower:
                output_ws.cell(row=row, column=1).value = "Art"

    for row in range(DATA_START_ROW, output_ws.max_row + 1):
        output_ws.cell(row=row, column=17).value = None

    seen = set()
    rows_to_delete = []

    for row in range(DATA_START_ROW, output_ws.max_row + 1):
        row_key = tuple(output_ws.cell(row=row, column=col).value for col in range(1, 18))
        if row_key in seen:
            rows_to_delete.append(row)
        else:
            seen.add(row_key)

    for row in reversed(rows_to_delete):
        output_ws.delete_rows(row)

    output_buffer = io.BytesIO()
    output_wb.save(output_buffer)
    output_buffer.seek(0)

    return output_buffer.getvalue()


def main():
    # Header
    st.markdown("""
    <div class="main-header">
        <div class="main-title">📊 TALIMEX Internal TSS Converter</div>
        <div class="main-subtitle">Convert TALIMEX Internal TSS to Standard Internal TSS</div>
    </div>
    <div class="divider"></div>
    """, unsafe_allow_html=True)

    # Upload section
    st.markdown("""
    <div class="upload-container">
        <div class="upload-title">📁 Upload Excel File</div>
        <div class="upload-subtitle">Select .xlsx file to convert</div>
    </div>
    """, unsafe_allow_html=True)

    # File uploader
    uploaded_files = st.file_uploader(
        "Upload files",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        label_visibility="collapsed"
    )

    if not uploaded_files:
        return

    st.markdown("---")

    # Validation
    st.markdown("#### Validation Results")

    validation_results = []
    for uploaded_file in uploaded_files:
        file_bytes = uploaded_file.read()
        uploaded_file.seek(0)
        result = validate_file_content(file_bytes, uploaded_file.name)
        validation_results.append((uploaded_file, result))

    valid_files = []
    invalid_files = []

    for uploaded_file, result in validation_results:
        if result.is_valid:
            valid_files.append(uploaded_file)
            st.markdown(f'<span class="file-valid">✓ {uploaded_file.name}</span> - Valid', unsafe_allow_html=True)
        else:
            invalid_files.append((uploaded_file, result))
            st.markdown(f'<span class="file-invalid">✗ {uploaded_file.name}</span> - Invalid', unsafe_allow_html=True)
            for error in result.errors:
                st.caption(f"   Column {error.column_letter}: expected '{error.expected}', got '{error.actual}'")

    if invalid_files:
        st.error(f"{len(invalid_files)} file(s) have invalid format. Please fix and re-upload.")
        return

    st.success(f"All {len(valid_files)} file(s) are valid and ready to process.")

    st.markdown("---")

    # Convert button
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        convert_clicked = st.button("🚀 Convert Files", type="primary", use_container_width=True)

    if convert_clicked:
        progress_bar = st.progress(0)
        status_text = st.empty()

        processed_files = []

        for uploaded_file in valid_files:
            status_text.text(f"Processing: {uploaded_file.name}")
            file_bytes = uploaded_file.read()

            def update_progress(step, status):
                progress_bar.progress(step / 4)
                status_text.text(f"Step {step}/4: {status}")

            try:
                result_bytes = process_file(file_bytes, uploaded_file.name, update_progress)
                input_name = Path(uploaded_file.name).stem
                output_name = f"{input_name}-Converted.xlsx"
                processed_files.append((output_name, result_bytes))

            except Exception as e:
                st.error(f"Error processing {uploaded_file.name}: {str(e)}")
                return

        progress_bar.progress(1.0)
        status_text.empty()
        st.success("All files processed successfully!")

        st.session_state['processed_files'] = processed_files

    # Download section
    if 'processed_files' in st.session_state and st.session_state['processed_files']:
        st.markdown("---")
        st.markdown("#### Download Results")

        processed_files = st.session_state['processed_files']

        if len(processed_files) == 1:
            filename, file_bytes = processed_files[0]
            st.download_button(
                label=f"📥 Download {filename}",
                data=file_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                for filename, file_bytes in processed_files:
                    zf.writestr(filename, file_bytes)
            zip_buffer.seek(0)

            st.download_button(
                label=f"📥 Download All ({len(processed_files)} files as ZIP)",
                data=zip_buffer.getvalue(),
                file_name="TSS_Converted_Files.zip",
                mime="application/zip",
                use_container_width=True
            )


if __name__ == "__main__":
    main()
