"""
Bước 2: Điền thông tin Product Name và Article Number
- Đọc file input để lấy product name và article number
- Mở file Step1 output và điền thông tin vào columns R onwards
- Xuất file: {input_name}-Step2.xlsx
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

INPUT_FOLDER = "input"
OUTPUT_FOLDER = "output"


def get_input_files() -> list[Path]:
    """Lấy tất cả file Excel từ folder input (bỏ qua file tạm ~$)"""
    input_path = Path(INPUT_FOLDER)
    files = []
    for pattern in ["*.xlsx", "*.xls"]:
        for f in input_path.glob(pattern):
            # Bỏ qua file tạm của Excel (bắt đầu bằng ~$)
            if not f.name.startswith("~$"):
                files.append(f)
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file Excel trong folder {INPUT_FOLDER}")
    return files


def get_sheets_except_material_code(wb):
    """Lấy tất cả sheet trừ 'material code' (case insensitive)"""
    sheets = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() != "material code":
            sheets.append(wb[sheet_name])
    return sheets


def find_product_info(ws):
    """Tìm Product name và Article number trong 3 row đầu"""
    product_names = []
    article_numbers = []

    for row in range(1, 4):  # Row 1-3
        for col in range(1, min(ws.max_column + 1, 20)):  # Giới hạn cột tìm kiếm
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_str = str(cell_value).lower().strip()
                if "product name" in cell_str:
                    # Tìm giá trị trong các cột tiếp theo (offset 1-3)
                    for offset in range(1, 4):
                        value = ws.cell(row=row, column=col + offset).value
                        if value:
                            product_names = [p.strip() for p in str(value).split('\n') if p.strip()]
                            break
                elif "article number" in cell_str:
                    # Tìm giá trị trong các cột tiếp theo (offset 1-3)
                    for offset in range(1, 4):
                        value = ws.cell(row=row, column=col + offset).value
                        if value:
                            article_numbers = [str(a).strip() for a in str(value).split('\n') if str(a).strip()]
                            break

    return product_names, article_numbers


def fill_product_columns(ws, product_names, article_numbers):
    """Điền product info vào columns R onwards"""
    start_col = 18  # Column R = 18

    # Số cột = max của 2 danh sách
    num_products = max(len(product_names), len(article_numbers))

    # Style cho header
    header_font = Font(bold=False)
    header_alignment = Alignment(
        textRotation=90,
        vertical='center',
        horizontal='center',
        wrap_text=True
    )
    article_alignment = Alignment(
        vertical='center',
        horizontal='center'
    )

    # Background màu peach/salmon (ARGB format)
    peach_fill = PatternFill(start_color="FFFCD5B4", end_color="FFFCD5B4", fill_type="solid")

    for i in range(num_products):
        col = start_col + i
        col_letter = ws.cell(row=1, column=col).column_letter

        # Lấy giá trị (có thể None nếu danh sách ngắn hơn)
        name = product_names[i] if i < len(product_names) else ""
        article = article_numbers[i] if i < len(article_numbers) else ""

        # Article number ở row 10
        article_cell = ws.cell(row=10, column=col, value=article)
        article_cell.alignment = article_alignment
        article_cell.font = header_font
        article_cell.fill = peach_fill

        # Product name: merge row 1 đến row 9, xoay 90 độ
        ws.merge_cells(start_row=1, start_column=col, end_row=9, end_column=col)
        name_cell = ws.cell(row=1, column=col, value=name)
        name_cell.alignment = header_alignment
        name_cell.font = header_font
        name_cell.fill = peach_fill

        # Apply fill cho tất cả cells trong merged range (rows 1-9)
        for row in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = peach_fill

        # Tự động tính width dựa trên độ dài article number
        article_len = len(str(article)) if article else 0
        width = max(article_len + 2, 10)  # Tối thiểu 10 để dễ đọc
        ws.column_dimensions[col_letter].width = width


def main():
    input_files = get_input_files()

    print(f"Tìm thấy {len(input_files)} file trong folder input")

    for input_file in input_files:
        input_name = input_file.stem
        step1_file = Path(OUTPUT_FOLDER) / f"{input_name}-Step1.xlsx"

        if not step1_file.exists():
            print(f"  ⚠ Bỏ qua {input_file.name}: Chưa có file Step1")
            continue

        # Đọc file input
        input_wb = load_workbook(input_file)
        input_sheets = get_sheets_except_material_code(input_wb)

        if not input_sheets:
            print(f"  ⚠ Bỏ qua {input_file.name}: Không tìm thấy sheet phù hợp")
            continue

        # Lấy thông tin product từ tất cả các sheet (trừ material code)
        all_product_names = []
        all_article_numbers = []

        for input_ws in input_sheets:
            product_names, article_numbers = find_product_info(input_ws)
            if product_names:
                all_product_names.extend(product_names)
            if article_numbers:
                all_article_numbers.extend(article_numbers)

        # Kiểm tra: nếu không tìm thấy CẢ product name VÀ article number thì báo lỗi và dừng
        if not all_product_names and not all_article_numbers:
            print(f"\n❌ LỖI: File '{input_file.name}' không có Product name và Article number!")
            print(f"   Vui lòng kiểm tra lại file input và đảm bảo có ít nhất 1 trong 2 thông tin.")
            print(f"\n⛔ Workflow đã dừng.")
            return  # Dừng workflow

        # Mở file Step1
        output_wb = load_workbook(step1_file)
        output_ws = output_wb.active

        # Điền thông tin
        fill_product_columns(output_ws, all_product_names, all_article_numbers)

        # Lưu file Step2
        output_filename = f"{input_name}-Step2.xlsx"
        output_path = Path(OUTPUT_FOLDER) / output_filename
        output_wb.save(output_path)

        print(f"  {input_file.name} → {output_filename}")
        if all_product_names and all_article_numbers:
            print(f"    Tìm thấy {max(len(all_product_names), len(all_article_numbers))} products: {', '.join(all_article_numbers)}")
        elif all_product_names:
            print(f"    Tìm thấy {len(all_product_names)} product names (không có article numbers)")
        elif all_article_numbers:
            print(f"    Tìm thấy {len(all_article_numbers)} article numbers (không có product names)")

    print(f"\nHoàn thành!")


if __name__ == "__main__":
    main()
