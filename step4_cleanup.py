"""
Bước 4: Clean up dữ liệu
- Clear cột K nếu H không phải "Test report"/"TR"
- Điền "Art" vào A nếu Q chứa "Article"
- Xóa nội dung cột Q
- Loại bỏ dòng trùng lặp (A-Q)
- Xuất file: {input_name}-Step4.xlsx
"""

from pathlib import Path
from openpyxl import load_workbook

OUTPUT_FOLDER = "output"
DATA_START_ROW = 11  # Data bắt đầu từ row 11


def get_step3_files() -> list[Path]:
    """Lấy tất cả file Step3 từ folder output (bỏ qua file tạm ~$)"""
    output_path = Path(OUTPUT_FOLDER)
    files = []
    for f in output_path.glob("*-Step3.xlsx"):
        if not f.name.startswith("~$"):
            files.append(f)
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file Step3 trong folder {OUTPUT_FOLDER}")
    return files


def clear_k_if_not_test_report(ws):
    """
    Nếu H không phải 'Test Report' hoặc 'TR' (case insensitive)
    thì clear nội dung cột K ở dòng đó
    """
    count = 0
    for row in range(DATA_START_ROW, ws.max_row + 1):
        h_val = ws.cell(row=row, column=8).value  # Column H = 8
        if h_val:
            h_lower = str(h_val).lower().strip()
            if h_lower not in ('test report', 'tr'):
                ws.cell(row=row, column=11).value = None  # Column K = 11
                count += 1
    return count


def fill_art_if_article_in_q(ws):
    """
    Nếu Q chứa 'Article' hoặc 'Art' (case insensitive)
    thì điền 'Art' vào cột A ở hàng đó
    """
    count = 0
    for row in range(DATA_START_ROW, ws.max_row + 1):
        q_val = ws.cell(row=row, column=17).value  # Column Q = 17
        if q_val:
            q_lower = str(q_val).lower()
            if 'article' in q_lower or 'art' in q_lower:
                ws.cell(row=row, column=1).value = "Art"  # Column A = 1
                count += 1
    return count


def clear_column_q(ws):
    """Xóa toàn bộ nội dung cột Q từ DATA_START_ROW"""
    count = 0
    for row in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=row, column=17).value:
            ws.cell(row=row, column=17).value = None
            count += 1
    return count


def remove_duplicate_rows(ws):
    """
    Loại bỏ các dòng có nội dung trùng từ cột A đến Q
    Giữ dòng đầu tiên, xóa các dòng trùng sau
    """
    seen = set()
    rows_to_delete = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        # Tạo tuple từ giá trị A-Q để check trùng
        row_key = tuple(
            ws.cell(row=row, column=col).value
            for col in range(1, 18)  # A=1 to Q=17
        )

        if row_key in seen:
            rows_to_delete.append(row)
        else:
            seen.add(row_key)

    # Xóa từ dưới lên để không ảnh hưởng index
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

    return len(rows_to_delete)


def main():
    step3_files = get_step3_files()

    print(f"Tìm thấy {len(step3_files)} file Step3 trong folder output")

    for step3_file in step3_files:
        # Lấy tên gốc (bỏ -Step3)
        base_name = step3_file.stem.replace("-Step3", "")

        # Mở file Step3
        wb = load_workbook(step3_file)
        ws = wb.active

        # 1. Clear K nếu H không phải Test report/TR
        k_cleared = clear_k_if_not_test_report(ws)

        # 2. Điền Art vào A nếu Q chứa Article
        art_filled = fill_art_if_article_in_q(ws)

        # 3. Xóa nội dung cột Q
        q_cleared = clear_column_q(ws)

        # 4. Loại bỏ dòng trùng lặp
        duplicates_removed = remove_duplicate_rows(ws)

        # Lưu file Step4
        output_filename = f"{base_name}-Step4.xlsx"
        output_path = Path(OUTPUT_FOLDER) / output_filename
        wb.save(output_path)

        print(f"  {step3_file.name} → {output_filename}")
        print(f"    - Cleared K: {k_cleared} rows")
        print(f"    - Art filled: {art_filled} rows")
        print(f"    - Q cleared: {q_cleared} cells")
        print(f"    - Duplicates removed: {duplicates_removed} rows")

    print(f"\nHoàn thành!")


if __name__ == "__main__":
    main()
