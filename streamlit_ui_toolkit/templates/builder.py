"""
Excel Template Builder - Create formatted Excel files from template config
Extracted from SEDO TSS Converter step3_template_creation.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Union, Optional
from .schema import TemplateConfig


class ExcelTemplateBuilder:
    """
    Build Excel files from template configurations

    Creates formatted Excel workbooks with styled headers, column widths,
    and other formatting based on TemplateConfig.
    """

    def __init__(self, template: TemplateConfig):
        """
        Initialize ExcelTemplateBuilder

        Args:
            template: TemplateConfig instance defining the template structure
        """
        self.template = template

    def create_workbook(self,
                       output_path: Union[str, Path],
                       sheet_name: str = "Sheet1") -> str:
        """
        Create Excel workbook from template

        Args:
            output_path: Path for output Excel file
            sheet_name: Name for the worksheet

        Returns:
            Path to created file (as string)

        Example:
            template = TemplateConfig.from_yaml("template.yaml")
            builder = ExcelTemplateBuilder(template)
            output = builder.create_workbook("output/template.xlsx")
        """
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Apply column configurations
        for col_idx, col_config in enumerate(self.template.columns, 1):
            # Write header text
            cell = ws.cell(self.template.header_row, col_idx, col_config.name)

            # Apply font styling
            cell.font = Font(
                bold=col_config.font_bold,
                color=col_config.font_color
            )

            # Apply background fill
            cell.fill = PatternFill(
                start_color=col_config.bg_color,
                end_color=col_config.bg_color,
                fill_type="solid"
            )

            # Apply alignment
            cell.alignment = Alignment(
                horizontal=col_config.horizontal_align,
                vertical=col_config.vertical_align,
                wrap_text=col_config.wrap_text
            )

            # Set column width
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = col_config.width

        # Freeze panes if specified
        if self.template.freeze_panes:
            ws.freeze_panes = self.template.freeze_panes

        # Enable auto-filter if specified
        if self.template.auto_filter:
            # Set auto-filter on header row
            last_col = get_column_letter(len(self.template.columns))
            filter_range = f"A{self.template.header_row}:{last_col}{self.template.header_row}"
            ws.auto_filter.ref = filter_range

        # Save workbook
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        return str(output_path)

    def add_data_rows(self,
                     workbook_path: Union[str, Path],
                     data: list[list],
                     start_row: Optional[int] = None) -> str:
        """
        Add data rows to existing workbook

        Args:
            workbook_path: Path to existing workbook
            data: List of rows, where each row is a list of values
            start_row: Starting row number (1-indexed). If None, starts after header row.

        Returns:
            Path to updated file

        Example:
            data = [
                ["Value1", "Value2", "Value3"],
                ["Value4", "Value5", "Value6"],
            ]
            builder.add_data_rows("output.xlsx", data)
        """
        wb = openpyxl.load_workbook(workbook_path)
        ws = wb.active

        # Determine start row
        if start_row is None:
            start_row = self.template.header_row + 1

        # Write data
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row_idx, col_idx, value)

        # Save workbook
        wb.save(str(workbook_path))
        return str(workbook_path)
